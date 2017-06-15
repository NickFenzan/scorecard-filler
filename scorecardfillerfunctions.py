import csv
import datetime
import openpyxl
def csvToDictList(file):
    with open(file) as csvfile:
        reader = csv.DictReader(csvfile)
        rows = [x for x in reader]
    return rows

def countNonConsultAppointmentTypes(appointments):
    def rowMatches(row):
        return not any(row['appttype'] in atype for atype in ['Free Evaluation', 'New Patient'])
    count = 0
    for x in [int(row['chckdin appt cnt']) for row in appointments if rowMatches(row)]:
        count = count + x
    return count

def countAppointmentsBySiteAndType(appointments, site, types):
    def rowMatches(row):
        return (row['svc dprtmnt'] == site.upper()  and any(row['appttype'] in atype for atype in types))
    count = 0
    for x in [int(row['chckdin appt cnt']) for row in appointments if rowMatches(row)]:
        count = count + x
    return count

def determineWeeklyDate(appointments):
    apptDate = datetime.datetime.strptime(appointments[0]['apptdate'], '%m/%d/%Y')
    idx = (apptDate.weekday() + 1) % 7
    return apptDate - datetime.timedelta(idx)

def findDateCell(sheet, date):
    for col in sheet.iter_cols(min_row=1, min_col=4, max_row=1):
        for cell in col:
            if(cell.value == date):
                return cell
    return None

def findApptLabelCell(sheet, apptLabel):
    for row in sheet.iter_rows(min_col=1, max_col=2, min_row=2):
        for cell in row:
            if(cell.value == apptLabel):
                return cell
    return None

def writeAppointments(scorecard, weeklyDate, sites, apptCategories, appointments):
    for site in sites:
        sheet = scorecard[site]
        dateCell = findDateCell(sheet, weeklyDate)
        for apptLabel, appttypes in apptCategories.items():
            apptLabelCell = findApptLabelCell(sheet, apptLabel)
            apptCount = countAppointmentsBySiteAndType(appointments, site, appttypes)
            sheet[dateCell.column+str(apptLabelCell.row)] = apptCount
    sheet = scorecard['Call Data']
    dateCell = findDateCell(sheet, weeklyDate)
    apptLabelCell = findApptLabelCell(sheet, "Other Appointments")
    otherAppointments = countNonConsultAppointmentTypes(appointments)
    sheet[dateCell.column+str(apptLabelCell.row)] = otherAppointments

def writeBilled(scorecard, weeklyDate, sites, billed):
    for site in sites:
        sheet = scorecard[site]
        dateCell = findDateCell(sheet, weeklyDate)
        billedAmt = next(float(x['all chgs']) for x in billed if x['svc dprtmnt'] == site.upper())
        apptLabelCell = findApptLabelCell(sheet, "Billed")
        sheet[dateCell.column+str(apptLabelCell.row)] = billedAmt

def writeCollected(scorecard, weeklyDate, sites, collected):
    for site in sites:
        sheet = scorecard[site]
        dateCell = findDateCell(sheet, weeklyDate)
        collectedAmt = next(float(x['payments']) for x in collected if x['svc dprtmnt'] == site.upper())
        apptLabelCell = findApptLabelCell(sheet, "Collected")
        sheet[dateCell.column+str(apptLabelCell.row)] = collectedAmt

def writeStaffing(scorecard, weeklyDate, sites, staffRoles, staffing):
    for site in sites:
        sheet = scorecard[site]
        dateCell = findDateCell(sheet, weeklyDate)
        for staff in staffRoles:
            apptLabelCell = findApptLabelCell(sheet, staff)
            staffCount = next((int(x['count']) for x in staffing if x['city'] == site and x['role'] == staff.lower()),0)
            sheet[dateCell.column+str(apptLabelCell.row)] = staffCount

def writeFormSubmissions(scorecard, weeklyDate, sites, websiteforms):
    def rowTest(row):
        convDate = datetime.datetime.strptime(row['Conversion Date'], '%Y-%m-%d %I:%M:%S %p')
        return (row['Location'] == site and convDate >= weeklyDate and convDate <= nextWeek)
    nextWeek = weeklyDate + datetime.timedelta(7)
    for site in sites:
        sheet = scorecard[site]
        dateCell = findDateCell(sheet, weeklyDate)
        apptLabelCell = findApptLabelCell(sheet, "Form Submissions")
        formCount = len([x for x in websiteforms if rowTest(x)])
        sheet[dateCell.column+str(apptLabelCell.row)] = formCount

def writeCalls(scorecard, weeklyDate, calls):
    sheet = scorecard['Call Data']
    dateCell = findDateCell(sheet, weeklyDate)
    apptLabelCell = findApptLabelCell(sheet, "Total Calls")
    sheet[dateCell.column+str(apptLabelCell.row)] = calls

def writeSessions(scorecard, weeklyDate, sessions):
    sheet = scorecard['Overall']
    dateCell = findDateCell(sheet, weeklyDate)
    apptLabelCell = findApptLabelCell(sheet, "Sessions")
    sheet[dateCell.column+str(apptLabelCell.row)] = sessions

def writeScorecard(sites, apptCategories, staffRoles, args):
    scorecard = openpyxl.load_workbook(filename = args.scorecard)
    appointments = csvToDictList(args.appointments)
    weeklyDate = determineWeeklyDate(appointments)
    billed = csvToDictList(args.billed)
    collected = csvToDictList(args.collected)
    staffing = csvToDictList(args.staffing)
    websiteforms = csvToDictList(args.websiteform)

    writeAppointments(scorecard, weeklyDate, sites, apptCategories, appointments)
    writeBilled(scorecard, weeklyDate, sites, billed)
    writeCollected(scorecard, weeklyDate, sites, collected)
    writeStaffing(scorecard, weeklyDate, sites, staffRoles, staffing)
    writeFormSubmissions(scorecard, weeklyDate, sites, websiteforms)
    writeCalls(scorecard, weeklyDate, args.calls)
    writeSessions(scorecard, weeklyDate, args.websessions)

    scorecard.save('scorecard.xlsx')
